from flask import Flask, request, send_file, render_template_string
import pdfplumber
import re
from openpyxl import Workbook
import logging
import io
import sys

app = Flask(__name__)
PDF_PASSWORD = '220165'  # You can make this dynamic if needed

# Logging setup
logging.basicConfig(stream=sys.stdout, level=logging.INFO,
                    format='%(asctime)s - %(message)s')

# HTML upload form
UPLOAD_FORM = """
<!doctype html>
<title>Upload PDF</title>
<h2>Upload Mutual Fund PDF</h2>
<form method=post enctype=multipart/form-data>
  <input type=file name=file>
  <input type=submit value=Upload>
</form>
"""

def clean_and_convert_to_number(value):
    if value is None:
        return None, False
    cleaned = re.sub(r'[^\d.]', '', str(value))
    try:
        return float(cleaned), True
    except (ValueError, TypeError):
        return value, False

def process_pdf(file_stream, ws, row_counter, password):
    Fund_Name = ""
    Folio_No = ""
    ISIN = ""
    i = row_counter

    try:
        with pdfplumber.open(file_stream, password=password) as invoice_pdf:
            for page in invoice_pdf.pages:
                for line in page.extract_text().splitlines():
                    purchase = re.search(r"Purchase.*|Redemption.*|\(\d+\/\d+", line)
                    folio = re.search(r"Folio No:", line)
                    fund = re.search(r"ISIN:", line)

                    if folio:
                        folio_cut = line.split()
                        try:
                            folio_index = folio_cut.index("No:") + 1
                            Folio_No = folio_cut[folio_index] if folio_index < len(folio_cut) else ""
                        except ValueError:
                            Folio_No = ""

                    if fund:
                        try:
                            isin_match = re.search(r"ISIN:\s*([A-Za-z0-9]{1,13})(?:\s|\W|$)", line)
                            ISIN = isin_match.group(1) if isin_match else ""
                        except:
                            ISIN = ""
                        try:
                            line_parts = line.split('-', 2)
                            Fund_Name = line_parts[1].strip() if len(line_parts) >= 2 else ""
                        except:
                            Fund_Name = ""

                    if purchase:
                        words = line.split()
                        if len(words) >= 4:
                            Amt, Unit, Nav, Unit_Bal = words[-4:]
                            Trans_date, Trans_type = words[0], words[1]

                            cleaned_amt, is_amt_numeric = clean_and_convert_to_number(Amt)
                            cleaned_unit, is_unit_numeric = clean_and_convert_to_number(Unit)
                            cleaned_nav, is_nav_numeric = clean_and_convert_to_number(Nav)
                            cleaned_unit_bal, is_unit_bal_numeric = clean_and_convert_to_number(Unit_Bal)

                            if all([is_amt_numeric, is_unit_numeric, is_nav_numeric, is_unit_bal_numeric]):
                                i += 1
                                ws[f'A{i+1}'] = Trans_date
                                ws[f'B{i+1}'] = Trans_type
                                ws[f'C{i+1}'] = Fund_Name
                                ws[f'D{i+1}'] = cleaned_amt
                                ws[f'E{i+1}'] = cleaned_unit
                                ws[f'F{i+1}'] = cleaned_nav
                                ws[f'G{i+1}'] = cleaned_unit_bal
                                ws[f'H{i+1}'] = ISIN
                                ws[f'I{i+1}'] = Folio_No
                            else:
                                logging.info(f"Removed row: Line='{line}', "
                                             f"Non-numeric values: Amt='{Amt}', Unit='{Unit}', "
                                             f"Nav='{Nav}', Unit_Bal='{Unit_Bal}'")
        return i
    except Exception as e:
        logging.error(f"Error processing PDF: {str(e)}")
        return row_counter

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        uploaded_file = request.files.get('file')
        if not uploaded_file or uploaded_file.filename == '':
            return "No file uploaded."

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        headers = ["Trans_date", "Trans_type", "Fund_Name", "Amount", "Units", "NAV", "Unit_Balance", "ISIN", "Folio"]
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)

        row_counter = process_pdf(uploaded_file.stream, ws, 0, PDF_PASSWORD)

        # Save to in-memory file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, download_name="output.xlsx", as_attachment=True)

    return render_template_string(UPLOAD_FORM)

if __name__ == '__main__':
    app.run(debug=True)